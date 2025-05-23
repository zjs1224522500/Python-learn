import os

import numpy as np
from matplotlib import pyplot as plt
from openpyxl import load_workbook


class YearData:
    def __init__(self, p_year, p_min_rank, p_avg_rank):
        self.year = p_year
        self.min = 0 if '--' == p_min_rank else p_min_rank
        self.avg = 0 if '--' == p_avg_rank else p_avg_rank


class SchoolData:
    def __init__(self, p_name, p_years):
        self.name = p_name
        self.years = p_years


path = os.getcwd() + r'\excel\university.xlsx'
print(path)
# exit()
"""读取excel文件,API见https://xlrd.readthedocs.io/en/latest/api.html"""
filename = path
book_wind = load_workbook(filename=filename)
book_sheet = book_wind.active

# 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
# sheets = workbook.get_sheet_names()  # 从名称获取sheet
# book_sheet = workbook.get_sheet_by_name(sheets[0])

# 获取sheet页的行数据
rows = book_sheet.rows
# 获取sheet页的列数据
columns = book_sheet.columns

i = 0
years = []
schools = []
for row in rows:
    i = i + 1
    line = [col.value for col in row][0:13]

    if i >= 3:
        school_name = book_sheet.cell(row=i, column=1).value  # 获取第i行1 列的数据
        col_count = 0

        year_list = []
        for year in years:
            col_count += 2
            min_rank = book_sheet.cell(row=i, column=col_count).value  # 获取第i行 2 列的数据
            avg_rank = book_sheet.cell(row=i, column=col_count + 1).value  # 获取第i行 3 列的数据
            year_data = YearData(year, min_rank, avg_rank)
            year_list.append(year_data)
        school_data = SchoolData(school_name, year_list)
        schools.append(school_data)
        # print(school_name, cell_data_2, cell_data_3, cell_data_4)
        # print(str(school_data.__dict__))
    if i == 1:
        years = line[1:13:2]

# 绘制曲线图
count = 0
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

i = 0
for i in range(2):
    print(i)
    plt.figure(figsize=(20, 10))
    for school_data in schools:
        ax = plt.gca()
        ax.invert_yaxis()

        count += 1
        min_data_arr = [year.min for year in school_data.years][::-1]
        avg_data_arr = [year.avg for year in school_data.years][::-1]
        data = min_data_arr if i == 0 else avg_data_arr
        label_name = str('min' + school_data.name) if i == 0 else str('avg' + school_data.name)
        plt.subplot(2, 4, int((count - 1) / 4 + 1))
        plt.plot(years[::-1], data, label=label_name)
        plt.legend(loc='best', fontsize='small')

        for a, b in zip(years[::-1], data):
            plt.text(a, b, b, ha='center', va='bottom', fontsize=10)

        # if count % 8 == 0:

        # line1.set_dashes([2, 2, 10, 2])  # 将曲线设置为点划线，set_dashes([line_space,space_space,line_space,space_space])
        # line2, = plt.plot(x, y2, label='流量场方差')
        # line2.set_dashes([2, 2, 2, 2])
        # plt.title('方差曲线', fontsize=16)
        # plt.legend(loc=4)  # 设置图例位置，4表示右下角
    count = 0

    ax = plt.gca()
    # # ax.xaxis.set_ticks_position('top')  # 将x轴的位置设置在顶部
    # # ax.invert_xaxis()  # x轴反向
    #
    # # ax.yaxis.set_ticks_position('right')  # 将y轴的位置设置在右边
    ax.invert_yaxis()  # y轴反向

    plt.show()

